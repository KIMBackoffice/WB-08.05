# src/export_zuweisung_xlsx.py
"""
Excel-Export für den Tab «Zuweisung» (manuelle Zuweisung).

Erzeugt eine flache Analyse-Tabelle: EINE ZEILE PRO ZUGEWIESENER PERSON.
Journal Club hat zwei Slots (OA/Int. + AA) → zwei Zeilen, Slot 1 und Slot 2.
Alle anderen Events haben nur Slot 1.

Pro Zeile:
  · Termin (Datum, Wochentag, Zeit von/bis, Plan-Code, Veranstaltung)
  · Zugewiesene Person + PEP-Rolle + Fellow/Rotation + Dienstcode + Dienstname
  · Letzte Zuweisung (aus History + bereits geplanten Monaten)
  · Nächste Zuweisung (aus dem Mehrmonatsplan)
  · Abstands-Checks (>30 Tage seit letzter / seit Arbeitsbeginn / bis nächster)
  · Bis zu 7 Alternativkandidaten mit Prio-Stufe, Rolle, Tagesdienst,
    letzter und nächster Zuweisung

Keine Google-Sheets-Zugriffe hier — alle Daten kommen als DataFrames rein.
"""
from __future__ import annotations

import io
import datetime
import pandas as pd

from src.fairness import (
    RELEVANT_EVENTS,
    EVENT_DUTY_RULES,
    _find_alternatives_ordered,
    _get_duty_priority_label,
)
from src.utils_names import extract_lastname, format_single_person


# ── Konstanten ─────────────────────────────────────────────────────────────

MAX_CANDIDATES = 7

WEEKDAY_DE_SHORT = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

_MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mär", 4: "Apr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Dez",
}

# Kurzcode wie im Plan
EVENT_CODE = {
    "COD_SENIOR":          "S-COD",
    "COD_JUNIOR":          "COD",
    "PEER":                "PEER",
    "PHYSIO":              "PHYSIO",
    "Mittwoch_Curriculum": "MI",
    "Journal_Club":        "JC",
}

EVENT_LABEL = {
    "COD_SENIOR":          "S - Case of the Day (COD)",
    "COD_JUNIOR":          "Case of the Day (COD)",
    "PEER":                "Peer-Teaching Session",
    "PHYSIO":              "Physiologie Talk",
    "Mittwoch_Curriculum": "Mittwochscurriculum",
    "Journal_Club":        "Journal Club",
}

# Slot-Bezeichnung je Event (Reihenfolge = Reihenfolge in EVENT_DUTY_RULES)
SLOT_LABEL = {
    "Journal_Club": ["OA / Int.", "AA"],
}

_GAP_DAYS = 30


# ── kleine Helfer ──────────────────────────────────────────────────────────

def _s(val, fallback: str = "") -> str:
    if val is None:
        return fallback
    if isinstance(val, float) and pd.isna(val):
        return fallback
    s = str(val).strip()
    return s if s and s.lower() != "nan" else fallback


def _fmt_date(d) -> str:
    if d is None or pd.isna(d):
        return ""
    return pd.Timestamp(d).strftime("%d.%m.%Y")


def _split_time(time_str: str) -> tuple[str, str]:
    """'11:30-11:45' → ('11:30', '11:45'). Ohne Bindestrich: alles in von."""
    t = _s(time_str)
    if not t:
        return "", ""
    for sep in ("-", "–", "—"):
        if sep in t:
            a, b = t.split(sep, 1)
            return a.strip(), b.strip()
    return t, ""


def _jn(flag) -> str:
    if flag is None:
        return "—"
    return "JA" if flag else "NEIN"


# ── Zuweisungs-Historie (History-Sheet + bereits geplante Monate) ───────────

def _build_assignment_log(history_df, schedule_all) -> pd.DataFrame:
    """
    Ein gemeinsames Log aller relevanten Zuweisungen.
    Spalten: lastname | date | event_type | topic | source
    Quellen: History-Sheet (Vergangenheit) + Mehrmonatsplan (Zukunft).
    """
    rows = []

    if history_df is not None and not history_df.empty:
        h = history_df.copy()
        if "date" in h.columns:
            h["date"] = pd.to_datetime(h["date"], errors="coerce", dayfirst=True)
        name_col = None
        for c in ("responsible_clean", "person", "responsible"):
            if c in h.columns:
                name_col = c
                break
        if name_col and "date" in h.columns:
            for _, r in h.iterrows():
                if pd.isna(r["date"]):
                    continue
                evt = _s(r.get("event_type"))
                if evt and evt not in RELEVANT_EVENTS:
                    continue
                for p in _s(r.get(name_col)).split("/"):
                    ln = extract_lastname(p.strip())
                    if not ln:
                        continue
                    rows.append({
                        "lastname":   ln,
                        "date":       pd.Timestamp(r["date"]).normalize(),
                        "event_type": evt,
                        "topic":      _s(r.get("topic")),
                        "source":     "History",
                    })

    if schedule_all is not None and not schedule_all.empty:
        s = schedule_all.copy()
        s["date"] = pd.to_datetime(s["date"], errors="coerce")
        s = s[s["event_type"].isin(RELEVANT_EVENTS)]
        for _, r in s.iterrows():
            if pd.isna(r["date"]):
                continue
            for p in _s(r.get("responsible")).split("/"):
                ln = extract_lastname(p.strip())
                if not ln:
                    continue
                rows.append({
                    "lastname":   ln,
                    "date":       pd.Timestamp(r["date"]).normalize(),
                    "event_type": _s(r.get("event_type")),
                    "topic":      _s(r.get("topic")),
                    "source":     "Plan",
                })

    if not rows:
        return pd.DataFrame(columns=["lastname", "date", "event_type", "topic", "source"])

    log = pd.DataFrame(rows).drop_duplicates(subset=["lastname", "date", "event_type"])
    return log.sort_values("date")


def _last_before(log: pd.DataFrame, lastname: str, ref_date) -> dict | None:
    if log.empty or not lastname:
        return None
    sub = log[(log["lastname"] == lastname) & (log["date"] < ref_date)]
    if sub.empty:
        return None
    return sub.iloc[-1].to_dict()


def _next_after(log: pd.DataFrame, lastname: str, ref_date) -> dict | None:
    if log.empty or not lastname:
        return None
    sub = log[(log["lastname"] == lastname) & (log["date"] > ref_date)]
    if sub.empty:
        return None
    return sub.iloc[0].to_dict()


# ── PEP-Helfer ─────────────────────────────────────────────────────────────

def _prepare_pep(pep_df) -> pd.DataFrame:
    p = pep_df.copy()
    p["date"]       = pd.to_datetime(p["date"], errors="coerce").dt.normalize()
    p["name_clean"] = p["name_clean"].astype(str).str.strip().str.lower()
    p["lastname"]   = p["name_clean"].apply(extract_lastname)
    p["duty_code"]  = pd.to_numeric(p["duty_code"], errors="coerce")
    p["role_code"]  = p["role_code"].astype(str).str.strip()
    return p


def _first_pep_date(pep_norm: pd.DataFrame, lastname: str):
    if not lastname:
        return None
    sub = pep_norm[pep_norm["lastname"] == lastname]
    if sub.empty:
        return None
    d = sub["date"].min()
    return None if pd.isna(d) else d


def _pep_day_row(pep_norm: pd.DataFrame, lastname: str, day) -> dict:
    sub = pep_norm[(pep_norm["lastname"] == lastname) & (pep_norm["date"] == day)]
    if sub.empty:
        return {}
    return sub.iloc[0].to_dict()


def _aa_type(aa_registry: dict, name_clean: str, role: str) -> str:
    """Fellow/Rotation — nur für AA sinnvoll."""
    if role != "AA":
        return ""
    if not aa_registry:
        return ""
    return _s(aa_registry.get(_s(name_clean).lower()))


# ── Kernaufbau ─────────────────────────────────────────────────────────────

def build_zuweisung_rows(
    schedule_month: pd.DataFrame,
    pep_df: pd.DataFrame,
    history_df=None,
    schedule_all=None,
    aa_registry: dict | None = None,
) -> pd.DataFrame:
    """Baut die flache Export-Tabelle (eine Zeile pro zugewiesene Person)."""
    pep_norm = _prepare_pep(pep_df) if pep_df is not None and not pep_df.empty else pd.DataFrame(
        columns=["date", "name_clean", "lastname", "duty_code", "role_code"]
    )
    log = _build_assignment_log(history_df, schedule_all)

    sc = schedule_month.copy()
    sc["date"] = pd.to_datetime(sc["date"], errors="coerce")
    sc = sc[sc["event_type"].isin(RELEVANT_EVENTS)].sort_values(["date", "time"])

    out_rows = []

    for _, row in sc.iterrows():
        evt   = _s(row.get("event_type"))
        day   = pd.Timestamp(row["date"]).normalize()
        t_von, t_bis = _split_time(row.get("time"))
        rules = EVENT_DUTY_RULES.get(evt, [])
        n_slots = max(1, len(rules))

        persons_raw = [p.strip() for p in _s(row.get("responsible")).split("/") if p.strip()]
        assigned_lns = [extract_lastname(p) for p in persons_raw]
        day_pep = pep_norm[pep_norm["date"] == day] if not pep_norm.empty else pd.DataFrame()

        for slot_idx in range(n_slots):
            person_raw = persons_raw[slot_idx] if slot_idx < len(persons_raw) else ""
            ln         = extract_lastname(person_raw)
            pep_row    = _pep_day_row(pep_norm, ln, day) if ln and not pep_norm.empty else {}
            role       = _s(pep_row.get("role_code"))
            duty       = pep_row.get("duty_code")
            name_clean = _s(pep_row.get("name_clean"), person_raw.lower())

            last_a = _last_before(log, ln, day)
            next_a = _next_after(log, ln, day)
            first_pep = _first_pep_date(pep_norm, ln) if ln and not pep_norm.empty else None

            days_since = (day - last_a["date"]).days if last_a else None
            days_until = (next_a["date"] - day).days if next_a else None
            days_start = (day - first_pep).days if first_pep is not None else None

            rec = {
                "monat":            _MONTH_NAMES.get(day.month, str(day.month)),
                "datum":            _fmt_date(day),
                "wochentag":        WEEKDAY_DE_SHORT[day.weekday()],
                "zeit_von":         t_von,
                "zeit_bis":         t_bis,
                "code":             EVENT_CODE.get(evt, evt),
                "veranstaltung":    EVENT_LABEL.get(evt, evt),
                "thema":            _s(row.get("topic")),
                "raum":             _s(row.get("room")),
                "slot":             slot_idx + 1,
                "slot_bezeichnung": (SLOT_LABEL.get(evt, [])[slot_idx]
                                     if slot_idx < len(SLOT_LABEL.get(evt, [])) else "—"),
                "person":           format_single_person(person_raw) if person_raw else "— tbd —",
                "person_pep":       name_clean,
                "rolle_pep":        role or "—",
                "funktion_detail":  _aa_type(aa_registry, name_clean, role) or "—",
                "dienst_code":      int(duty) if pd.notna(duty) else "",
                "dienst_bezeichnung": _get_duty_priority_label(duty, evt, role) if pd.notna(duty) else "—",

                "letzte_zuw_datum":   _fmt_date(last_a["date"]) if last_a else "",
                "letzte_zuw_event":   EVENT_CODE.get(last_a["event_type"], last_a["event_type"]) if last_a else "",
                "letzte_zuw_thema":   _s(last_a["topic"]) if last_a else "",
                "letzte_zuw_quelle":  last_a["source"] if last_a else "",
                "tage_seit_letzter":  days_since if days_since is not None else "",
                "check_30d_letzte":   _jn(days_since > _GAP_DAYS) if days_since is not None else "—",

                "naechste_zuw":       _jn(next_a is not None),
                "naechste_zuw_datum": _fmt_date(next_a["date"]) if next_a else "",
                "naechste_zuw_event": EVENT_CODE.get(next_a["event_type"], next_a["event_type"]) if next_a else "",
                "naechste_zuw_thema": _s(next_a["topic"]) if next_a else "",
                "tage_bis_naechster": days_until if days_until is not None else "",
                "check_30d_naechste": _jn(days_until > _GAP_DAYS) if days_until is not None else "—",

                "erster_pep_eintrag":      _fmt_date(first_pep) if first_pep is not None else "",
                "tage_seit_arbeitsbeginn": days_start if days_start is not None else "",
                "check_30d_arbeitsbeginn": _jn(days_start > _GAP_DAYS) if days_start is not None else "—",
            }

            # ── Alternativkandidaten ──────────────────────────────────────
            alts = []
            if slot_idx < len(rules) and not day_pep.empty:
                role_pool, duty_priority = rules[slot_idx]
                alts = _find_alternatives_ordered(
                    day_pep, role_pool, duty_priority, assigned_lns, event_date=day
                )

            for k in range(MAX_CANDIDATES):
                pfx = f"k{k+1}_"
                if k < len(alts):
                    a     = alts[k]
                    a_ln  = extract_lastname(a["name"])
                    a_last = _last_before(log, a_ln, day)
                    a_next = _next_after(log, a_ln, day)
                    rec[pfx + "prio"]     = a["priority_tier"]
                    rec[pfx + "name"]     = format_single_person(a["name"])
                    rec[pfx + "rolle"]    = a["role"]
                    rec[pfx + "dienst"]   = a["duty_label"]
                    rec[pfx + "letzte"]   = _fmt_date(a_last["date"]) if a_last else ""
                    rec[pfx + "naechste"] = _fmt_date(a_next["date"]) if a_next else ""
                else:
                    for suf in ("prio", "name", "rolle", "dienst", "letzte", "naechste"):
                        rec[pfx + suf] = ""

            out_rows.append(rec)

    return pd.DataFrame(out_rows)


# ── Excel-Ausgabe ──────────────────────────────────────────────────────────

_HEADERS_DE = {
    "monat": "Monat",
    "datum": "Datum", "wochentag": "Tag", "zeit_von": "Zeit von", "zeit_bis": "Zeit bis",
    "code": "Code", "veranstaltung": "Veranstaltung", "thema": "Thema", "raum": "Raum",
    "slot": "Slot", "slot_bezeichnung": "Slot-Typ",
    "person": "Person", "person_pep": "Person (PEP)", "rolle_pep": "Rolle",
    "funktion_detail": "Fellow/Rotation", "dienst_code": "Dienstcode",
    "dienst_bezeichnung": "Dienst",
    "letzte_zuw_datum": "Letzte Zuw. Datum", "letzte_zuw_event": "Letzte Zuw. Event",
    "letzte_zuw_thema": "Letzte Zuw. Thema", "letzte_zuw_quelle": "Quelle",
    "tage_seit_letzter": "Tage seit letzter", "check_30d_letzte": ">30d seit letzter",
    "naechste_zuw": "Nächste Zuw.", "naechste_zuw_datum": "Nächste Zuw. Datum",
    "naechste_zuw_event": "Nächste Zuw. Event", "naechste_zuw_thema": "Nächste Zuw. Thema",
    "tage_bis_naechster": "Tage bis nächster", "check_30d_naechste": ">30d bis nächster",
    "erster_pep_eintrag": "1. PEP-Eintrag", "tage_seit_arbeitsbeginn": "Tage seit Start",
    "check_30d_arbeitsbeginn": ">30d seit Start",
}

for _k in range(1, MAX_CANDIDATES + 1):
    _HEADERS_DE[f"k{_k}_prio"]     = f"K{_k} Prio"
    _HEADERS_DE[f"k{_k}_name"]     = f"K{_k} Name"
    _HEADERS_DE[f"k{_k}_rolle"]    = f"K{_k} Rolle"
    _HEADERS_DE[f"k{_k}_dienst"]   = f"K{_k} Dienst"
    _HEADERS_DE[f"k{_k}_letzte"]   = f"K{_k} letzte"
    _HEADERS_DE[f"k{_k}_naechste"] = f"K{_k} nächste"


def build_zuweisung_xlsx(
    schedule_month: pd.DataFrame,
    pep_df: pd.DataFrame,
    history_df=None,
    schedule_all=None,
    aa_registry: dict | None = None,
    month_label: str = "",
) -> bytes:
    """Baut die Tabelle und gibt formatierte .xlsx-Bytes zurück."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = build_zuweisung_rows(schedule_month, pep_df, history_df, schedule_all, aa_registry)

    wb = Workbook()
    ws = wb.active
    ws.title = "Zuweisung"

    if df.empty:
        ws.cell(row=1, column=1, value="Keine algorithmischen Veranstaltungen gefunden.")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    cols    = list(df.columns)
    headers = [_HEADERS_DE.get(c, c) for c in cols]

    hdr_fill_base = PatternFill("solid", fgColor="1A6E50")
    hdr_fill_cand = PatternFill("solid", fgColor="4A7BA7")
    hdr_font      = Font(bold=True, color="FFFFFF", size=10)
    thin          = Side(style="thin", color="D9D9D9")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_warn     = PatternFill("solid", fgColor="FFF3CD")
    fill_alt      = PatternFill("solid", fgColor="F5F8FA")

    for j, (col, head) in enumerate(zip(cols, headers), start=1):
        c = ws.cell(row=1, column=j, value=head)
        c.fill   = hdr_fill_cand if col.startswith("k") and col[1].isdigit() else hdr_fill_base
        c.font   = hdr_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    check_cols = {"check_30d_letzte", "check_30d_naechste", "check_30d_arbeitsbeginn"}

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        cand_row = (i % 2 == 0)
        for j, col in enumerate(cols, start=1):
            val = r[col]
            c = ws.cell(row=i, column=j, value=(None if val == "" else val))
            c.font   = Font(size=9.5)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(col in ("thema", "letzte_zuw_thema",
                                                                      "naechste_zuw_thema")))
            if col in check_cols and val == "NEIN":
                c.fill = fill_warn
                c.font = Font(size=9.5, bold=True, color="9C5700")
            elif cand_row and col.startswith("k") and col[1].isdigit():
                c.fill = fill_alt

    widths = {
        "monat": 8,
        "datum": 11, "wochentag": 6, "zeit_von": 9, "zeit_bis": 9, "code": 8,
        "veranstaltung": 24, "thema": 34, "raum": 12, "slot": 6, "slot_bezeichnung": 11,
        "person": 20, "person_pep": 20, "rolle_pep": 8, "funktion_detail": 13,
        "dienst_code": 10, "dienst_bezeichnung": 18,
        "letzte_zuw_datum": 13, "letzte_zuw_event": 13, "letzte_zuw_thema": 28,
        "letzte_zuw_quelle": 9, "tage_seit_letzter": 11, "check_30d_letzte": 13,
        "naechste_zuw": 10, "naechste_zuw_datum": 14, "naechste_zuw_event": 14,
        "naechste_zuw_thema": 28, "tage_bis_naechster": 12, "check_30d_naechste": 14,
        "erster_pep_eintrag": 13, "tage_seit_arbeitsbeginn": 12, "check_30d_arbeitsbeginn": 14,
    }
    for j, col in enumerate(cols, start=1):
        if col.startswith("k") and col[1].isdigit():
            suf = col.split("_", 1)[1]
            w   = {"prio": 7, "name": 18, "rolle": 8, "dienst": 17,
                   "letzte": 11, "naechste": 11}.get(suf, 12)
        else:
            w = widths.get(col, 14)
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    # ── Legende ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Legende")
    stamp = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    legend = [
        ("Export", f"Zuweisung {month_label}".strip()),
        ("Erstellt", stamp),
        ("", ""),
        ("Umfang", "Alle algorithmisch zugewiesenen Veranstaltungen der gewählten Monate."),
        ("Zeile", "Eine Zeile pro zugewiesene Person. Journal Club = 2 Zeilen (Slot 1 = OA/Int., Slot 2 = AA)."),
        ("Code", "S-COD, COD, PEER, PHYSIO, MI (Mittwochscurriculum), JC (Journal Club)"),
        ("Rolle", "Rollencode aus PEP am Veranstaltungstag (AA, OA_I, OA_II, SOA, SFA_II ...)"),
        ("Fellow/Rotation", "Aus dem AA-Registry-Sheet; nur für AA befüllt."),
        ("Dienst", "Dienstcode und Dienstbezeichnung aus PEP am Veranstaltungstag."),
        ("Letzte Zuw.", "Letzte Zuweisung vor diesem Datum — aus History-Sheet oder bereits geplanten Monaten (Spalte Quelle)."),
        ("Nächste Zuw.", "Nächste Zuweisung nach diesem Datum aus dem Mehrmonatsplan. Nur so weit, wie PEP-Daten vorhanden sind."),
        (">30d-Checks", "JA = Abstand grösser als 30 Tage (ok). NEIN = zu knapp (gelb). — = keine Vergleichsdaten."),
        ("K1–K7", "Alternativkandidaten in Prio-Reihenfolge (Prio 1 = beste Dienstgruppe), inkl. deren letzter/nächster Zuweisung."),
        ("Hinweis", "Nur algorithmisch zugewiesene Events (COD, PEER, PHYSIO, MI, JC). Sheet-basierte Events sind nicht enthalten."),
    ]
    for i, (k, v) in enumerate(legend, start=1):
        a = ws2.cell(row=i, column=1, value=k)
        a.font = Font(bold=True, size=10)
        b = ws2.cell(row=i, column=2, value=v)
        b.font = Font(size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
